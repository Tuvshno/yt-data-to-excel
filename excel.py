import os
import google_auth_oauthlib.flow
import googleapiclient.discovery
import googleapiclient.errors
from openpyxl import Workbook
from datetime import datetime

# Your channel ID
channel_id = "UCbz_aqiqO2HcAx01iJM5WGA"

# OAuth2 client secret file
client_secrets_file = "D:\Dev\yt-excel\client_secret_1091476056770-gif497vdakrujbodgd9apc9rb9hht84t.apps.googleusercontent.com.json"  # Replace with the path to your client_secret.json file


# Date filter (ISO 8601 format, e.g., '2022-01-01T00:00:00Z')
date_filter = "YYYY-MM-DDTHH:MM:SSZ"

scopes = ["https://www.googleapis.com/auth/youtube.force-ssl"]

# Get credentials and create an API client
flow = google_auth_oauthlib.flow.InstalledAppFlow.from_client_secrets_file(client_secrets_file, scopes)
credentials = flow.run_console()
youtube = googleapiclient.discovery.build("youtube", "v3", credentials=credentials)

def get_channel_videos(channel_id, date_filter):
    videos = []
    next_page_token = None

    while True:
        request = youtube.search().list(
            part="snippet",
            channelId=channel_id,
            maxResults=50,
            pageToken=next_page_token,
            type="video",
            publishedBefore=date_filter,
            order="date",
            fields="nextPageToken,items(id(videoId),snippet(publishedAt,title))"
        )
        response = request.execute()
        videos.extend(response["items"])
        next_page_token = response.get("nextPageToken")

        if next_page_token is None:
            break

    return videos

def get_video_details(video_ids):
    details = []
    for i in range(0, len(video_ids), 50):
        request = youtube.videos().list(
            part="snippet,statistics",
            id=','.join(video_ids[i:i + 50]),
            fields="items(id,snippet(publishedAt,title),statistics(viewCount))"
        )
        response = request.execute()
        details.extend(response["items"])

    return details

def save_to_excel(videos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Channel Videos"

    ws.cell(row=1, column=1, value="Title")
    ws.cell(row=1, column=2, value="Views")
    ws.cell(row=1, column=3, value="Video URL")
    ws.cell(row=1, column=4, value="Upload Date")

    for idx, video in enumerate(videos, start=2):
        ws.cell(row=idx, column=1, value=video["snippet"]["title"])
        ws.cell(row=idx, column=2, value=video["statistics"]["viewCount"])
        ws.cell(row=idx, column=3, value=f"https://www.youtube.com/watch?v={video['id']}")
        ws.cell(row=idx, column=4, value=video["snippet"]["publishedAt"])

    wb.save("channel_videos.xlsx")

channel_videos = get_channel_videos(channel_id, date_filter)
video_ids = [video["id"]["videoId"] for video in channel_videos]
videos_details = get_video_details(video_ids)
save_to_excel(videos_details)
